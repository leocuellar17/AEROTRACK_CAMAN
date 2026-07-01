# -*- coding: utf-8 -*-
"""
Módulo de Reportes - AeroTrack CAMAN
Maneja la generación de PDFs y exportación a Excel.
"""

import os
from datetime import datetime

# Importaciones para automatización de documentos
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_INSTALLED = True
except ImportError:
    REPORTLAB_INSTALLED = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_INSTALLED = True
except ImportError:
    OPENPYXL_INSTALLED = False

from database import DB_NAME, EXCEL_FILE, obtener_conexion

# Constantes Estéticas
COLOR_PRIMARY = "#0F2A4A"
COLOR_SECONDARY = "#1E5A99"
COLOR_ALERT = "#A83232"
COLOR_WARNING_BG = "#FFD6D6"
COLOR_WARNING_FG = "#9C0006"

def _aplicar_estilo_excel(ws, ultima_fila, fila_len):
    """
    Función helper para aplicar estilos a las filas de datos en Excel y autoajustar columnas.
    """
    font_data = Font(name="Calibri", size=10)
    border_thin = Side(border_style="thin", color="D3D3D3")
    border_box = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    fill_alerta = PatternFill(start_color=COLOR_WARNING_BG[1:], end_color=COLOR_WARNING_BG[1:], fill_type="solid")
    font_alerta = Font(name="Calibri", size=10, color=COLOR_WARNING_FG[1:], bold=True)
    
    for col_idx in range(1, fila_len + 1):
        cell = ws.cell(row=ultima_fila, column=col_idx)
        cell.font = font_data
        cell.border = border_box
        
        # Alineaciones específicas
        if col_idx in [1, 3, 5, 6, 7, 8]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
            
        # Formato de números para las horas
        if col_idx in [5, 6]:
            cell.number_format = '0.0'
            
        # Destacar celda de Estado en rojo preventivo
        if col_idx == 8:
            cell.fill = fill_alerta
            cell.font = font_alerta

    ws.row_dimensions[ultima_fila].height = 20
    
    # Ajustar automáticamente el ancho de las columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def generar_orden_trabajo(datos):
    """
    Genera una Orden de Trabajo oficial en PDF simulando el formato de CAMAN de la Fuerza Aérea Colombiana.
    """
    if not REPORTLAB_INSTALLED:
        return False, "ReportLab no está instalado en el sistema."

    id_componente = datos.get("id_componente")
    pdf_filename = f"OT_{id_componente}.pdf"
    
    try:
        doc = SimpleDocTemplate(
            pdf_filename,
            pagesize=letter,
            rightMargin=54, leftMargin=54, topMargin=54, bottomMargin=54
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        color_azul_caman = colors.HexColor(COLOR_PRIMARY)
        color_gris_suave = colors.HexColor("#f4f6f9")
        color_rojo_alerta = colors.HexColor(COLOR_ALERT)
        
        style_titulo = ParagraphStyle(
            'TituloCAMAN',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=color_azul_caman,
            alignment=1,
            spaceAfter=4
        )
        
        style_subtitulo = ParagraphStyle(
            'SubtituloCAMAN',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.HexColor("#555555"),
            alignment=1,
            spaceAfter=15
        )
        
        style_seccion = ParagraphStyle(
            'SeccionCAMAN',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=11,
            textColor=color_azul_caman,
            spaceBefore=10,
            spaceAfter=6
        )
        
        style_label = ParagraphStyle(
            'LabelCAMAN',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.HexColor("#333333")
        )
        
        style_valor = ParagraphStyle(
            'ValorCAMAN',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.HexColor("#222222")
        )
        
        style_alerta = ParagraphStyle(
            'AlertaCAMAN',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=color_rojo_alerta,
            alignment=1
        )
        
        # 1. ENCABEZADO OFICIAL
        bandera_data = [['', '', '']]
        bandera_table = Table(bandera_data, colWidths=[2.2*inch, 2.2*inch, 2.8*inch], rowHeights=[4])
        bandera_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), colors.HexColor("#FCD116")),
            ('BACKGROUND', (1,0), (1,0), colors.HexColor("#003893")),
            ('BACKGROUND', (2,0), (2,0), colors.HexColor("#CE1126")),
        ]))
        story.append(bandera_table)
        story.append(Spacer(1, 10))
        
        story.append(Paragraph("FUERZA AEROESPACIAL COLOMBIANA", style_titulo))
        story.append(Paragraph("COMANDO AÉREO DE MANTENIMIENTO - CAMAN", style_subtitulo))
        story.append(Paragraph("ORDEN DE TRABAJO DE MANTENIMIENTO", ParagraphStyle('DocTitle', parent=style_titulo, fontSize=12, spaceAfter=20)))
        
        # 2. INFORMACIÓN GENERAL
        info_data = [
            [Paragraph("Código de Orden:", style_label), Paragraph(f"OT-{id_componente}-{datetime.now().strftime('%Y%m%d')}", style_valor),
             Paragraph("Fecha de Emisión:", style_label), Paragraph(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), style_valor)],
            [Paragraph("Aeronave ID:", style_label), Paragraph(str(datos.get('id_avion')), style_valor),
             Paragraph("Modelo Aeronave:", style_label), Paragraph(str(datos.get('modelo_avion')), style_valor)],
            [Paragraph("Horas Totales Avión:", style_label), Paragraph(f"{datos.get('horas_totales_planeador'):.1f} Horas", style_valor),
             Paragraph("Estado de Alerta:", style_label), Paragraph("CRÍTICO (>=90%)", ParagraphStyle('EstadoCritico', parent=style_alerta, alignment=0))]
        ]
        
        info_table = Table(info_data, colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), color_gris_suave),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cccccc")),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 15))
        
        # 3. DETALLE COMPONENTE
        story.append(Paragraph("DETALLES DEL COMPONENTE CRÍTICO", style_seccion))
        
        horas_act = datos.get('horas_actuales')
        lim_horas = datos.get('limite_horas')
        porcentaje_uso = (horas_act / lim_horas) * 100 if lim_horas > 0 else 0.0
        
        comp_data = [
            [Paragraph("ID Componente", style_label), Paragraph("Nombre de Componente", style_label), 
             Paragraph("Horas Actuales", style_label), Paragraph("Límite de Vida", style_label), Paragraph("% de Ciclo", style_label)],
            [Paragraph(id_componente, style_valor), Paragraph(datos.get('nombre'), style_valor), 
             Paragraph(f"{horas_act:.1f} h", style_valor), Paragraph(f"{lim_horas:.1f} h", style_valor), 
             Paragraph(f"{porcentaje_uso:.1f}%", ParagraphStyle('PorcentajeColor', parent=style_alerta, fontSize=9))]
        ]
        
        comp_table = Table(comp_data, colWidths=[1.2*inch, 2.8*inch, 1.1*inch, 1.1*inch, 1.0*inch])
        comp_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), color_azul_caman),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#b2c2d4")),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 8),
        ]))
        for col_idx in range(len(comp_data[0])):
            comp_table.setStyle(TableStyle([
                ('TEXTCOLOR', (col_idx,0), (col_idx,0), colors.white),
            ]))
        story.append(comp_table)
        story.append(Spacer(1, 20))
        
        # 4. ADVERTENCIA
        advertencia_texto = (
            f"<b>ATENCIÓN:</b> El componente [{id_componente}] ha acumulado {horas_act:.1f} horas de "
            f"operación de un límite estructural/preventivo establecido de {lim_horas:.1f} horas. "
            f"Se ha alcanzado el <b>{porcentaje_uso:.1f}%</b> de su vida útil. Se ordena de forma inmediata la detención "
            f"preventiva de vuelos para la aeronave {datos.get('id_avion')} y el reemplazo/mantenimiento de dicho componente."
        )
        
        alerta_data = [[Paragraph(advertencia_texto, ParagraphStyle('AlertaTexto', parent=styles['Normal'], fontSize=9.5, textColor=color_rojo_alerta, leadingInterval=14))]]
        alerta_table = Table(alerta_data, colWidths=[7.2*inch])
        alerta_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#fff0f0")),
            ('BORDER', (0,0), (-1,-1), 1.5, color_rojo_alerta),
            ('PADDING', (0,0), (-1,-1), 12),
        ]))
        story.append(alerta_table)
        story.append(Spacer(1, 50))
        
        # 5. FIRMAS
        firma_data = [
            [Paragraph("<b>Emitido por:</b><br/>AeroTrack-CAMAN Logic Engine", style_valor), 
             Paragraph("<b>Firma Técnico Encargado:</b><br/><br/>________________________________________<br/>Tgo. de Mantenimiento CAMAN", style_valor)]
        ]
        firma_table = Table(firma_data, colWidths=[3.5*inch, 3.7*inch])
        firma_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('PADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(firma_table)
        
        doc.build(story)
        return True, None
        
    except PermissionError:
        return False, f"El archivo 'OT_{id_componente}.pdf' está abierto en otra aplicación. Por favor, ciérrelo."
    except Exception as e:
        return False, f"Error al generar PDF: {str(e)}"

def exportar_requerimiento_compras(datos):
    """
    Registra o actualiza el requerimiento de compras en el archivo Excel.
    """
    if not OPENPYXL_INSTALLED:
        return False, "La librería openpyxl no está instalada."

    id_componente = datos.get("id_componente")
    nombre_componente = datos.get("nombre")
    id_avion = datos.get("id_avion")
    modelo_avion = datos.get("modelo_avion")
    fecha_reporte = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    horas_actuales = datos.get("horas_actuales")
    limite_horas = datos.get("limite_horas")
    
    try:
        if os.path.exists(EXCEL_FILE):
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Requerimientos CAMAN"
            
            headers = ["ID Componente", "Nombre Componente", "Aeronave ID", "Modelo Aeronave", "Horas Actuales", "Límite Horas", "Fecha de Alerta", "Estado"]
            ws.append(headers)
            
            fill_header = PatternFill(start_color=COLOR_PRIMARY[1:], end_color=COLOR_PRIMARY[1:], fill_type="solid")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = align_header
                
            ws.row_dimensions[1].height = 28
            
        nueva_fila = [
            id_componente,
            nombre_componente,
            id_avion,
            modelo_avion,
            horas_actuales,
            limite_horas,
            fecha_reporte,
            "PENDIENTE ADQUISICIÓN"
        ]
        ws.append(nueva_fila)
        
        ultima_fila = ws.max_row
        _aplicar_estilo_excel(ws, ultima_fila, len(nueva_fila))
            
        wb.save(EXCEL_FILE)
        return True, None
        
    except PermissionError:
        return False, f"El archivo Excel '{EXCEL_FILE}' está abierto en otra aplicación. Por favor, ciérrelo."
    except Exception as e:
        return False, f"Error al generar Excel: {str(e)}"

def exportar_todos_los_criticos():
    """
    Genera el archivo excel conteniendo TODOS los componentes críticos.
    """
    if not OPENPYXL_INSTALLED:
        return 0, "La librería openpyxl no está instalada."
        
    conn = obtener_conexion()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT c.id_componente, c.nombre, c.id_avion, a.modelo, c.horas_actuales, c.limite_horas
        FROM Componentes c
        JOIN Aeronaves a ON c.id_avion = a.id_avion
        WHERE c.horas_actuales >= (c.limite_horas * 0.9);
    """)
    criticos = cursor.fetchall()
    conn.close()
    
    if not criticos:
        return 0, None
        
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requerimientos Críticos"
        
        headers = ["ID Componente", "Nombre Componente", "Aeronave ID", "Modelo Aeronave", "Horas Actuales", "Límite Horas", "Fecha de Alerta", "Estado"]
        ws.append(headers)
        
        fill_header = PatternFill(start_color=COLOR_PRIMARY[1:], end_color=COLOR_PRIMARY[1:], fill_type="solid")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_header
            
        ws.row_dimensions[1].height = 28
        
        fecha_reporte = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for idx, comp in enumerate(criticos):
            id_comp, nombre, id_avion, modelo, horas, limite = comp
            fila = [
                id_comp,
                nombre,
                id_avion,
                modelo,
                horas,
                limite,
                fecha_reporte,
                "PENDIENTE ADQUISICIÓN"
            ]
            ws.append(fila)
            row_idx = idx + 2
            
            _aplicar_estilo_excel(ws, row_idx, len(fila))
            
        wb.save(EXCEL_FILE)
        return len(criticos), None
        
    except PermissionError:
        return 0, f"El archivo Excel '{EXCEL_FILE}' está abierto en otra aplicación. Por favor, ciérrelo e intente de nuevo."
    except Exception as e:
        return 0, str(e)
